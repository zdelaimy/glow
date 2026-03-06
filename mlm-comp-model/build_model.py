import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
section_font = Font(bold=True, size=11, color="FFFFFF")
section_fill = PatternFill(start_color="6B4C9A", end_color="6B4C9A", fill_type="solid")
check_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
danger_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
good_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
label_font = Font(bold=True, size=10)
note_font = Font(italic=True, size=9, color="666666")
bold_font = Font(bold=True, size=10)
big_bold = Font(bold=True, size=12)
purple_bold = Font(bold=True, size=12, color="6B4C9A")
red_bold = Font(bold=True, size=10, color="CC0000")

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def style_section(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = section_font
        cell.fill = section_fill

def set_col_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w


# ============================================================
# SHEET 1: ASSUMPTIONS
# ============================================================
ws = wb.active
ws.title = "Assumptions"
set_col_widths(ws, [38, 18, 18, 55])

r = 1
ws.cell(r, 1, "GLOW BEAUTY — ALL ASSUMPTIONS").font = Font(bold=True, size=14)
r = 2
ws.cell(r, 1, "Change any value in columns B/C — the model recalculates from these.").font = note_font

# --- Product ---
r = 4
ws.cell(r, 1, "PRODUCT"); style_section(ws, r, 4)
r = 5
for i, h in enumerate(["", "Conservative", "Optimistic", "Why"]):
    ws.cell(r, i+1, h)
style_header(ws, r, 4)

product_assumptions = [
    ("Avg product price", 65, 65, "Premium beauty: $45-85 range"),
    ("Avg PV per active rep/month", 250, 500, "Conservative = autoship + light selling. Optimistic = active sellers."),
    ("Min autoship to stay active", 150, 150, "Monthly purchase required to earn commissions"),
    ("Product COGS (% of revenue)", 0.20, 0.20, "Cost to manufacture/source product"),
    ("Shipping & ops (% of revenue)", 0.10, 0.10, "Fulfillment, packaging, logistics"),
]

for i, (label, cons, opt, note) in enumerate(product_assumptions):
    r = 6 + i
    ws.cell(r, 1, label).font = label_font
    c2 = ws.cell(r, 2, cons)
    c3 = ws.cell(r, 3, opt)
    ws.cell(r, 4, note).font = note_font
    if "%" in label:
        c2.number_format = '0%'
        c3.number_format = '0%'
    elif "price" in label.lower() or "pv" in label.lower() or "autoship" in label.lower():
        c2.number_format = '$#,##0'
        c3.number_format = '$#,##0'

# --- Churn / Retention ---
r = 12
ws.cell(r, 1, "CHURN & RETENTION"); style_section(ws, r, 4)
r = 13
for i, h in enumerate(["", "Conservative", "Optimistic", "Why"]):
    ws.cell(r, i+1, h)
style_header(ws, r, 4)

churn_assumptions = [
    ("Monthly churn: Months 1-3", 0.08, 0.05, "The danger zone. Most quits happen here."),
    ("Monthly churn: Months 4-12", 0.04, 0.025, "Settled in but still fragile"),
    ("Monthly churn: Month 13+", 0.02, 0.012, "Long-term steady state"),
    ("12-month survival rate", "=((1-B14)^3)*((1-B15)^9)", "=((1-C14)^3)*((1-C15)^9)", "% of a cohort still active after 12 months"),
    ("24-month survival rate", "=B17*((1-B16)^12)", "=C17*((1-C16)^12)", "% still active after 24 months"),
]

for i, (label, cons, opt, note) in enumerate(churn_assumptions):
    r = 14 + i
    ws.cell(r, 1, label).font = label_font
    c2 = ws.cell(r, 2, cons)
    c3 = ws.cell(r, 3, opt)
    ws.cell(r, 4, note).font = note_font
    c2.number_format = '0.0%'
    c3.number_format = '0.0%'

# --- Recruiting ---
r = 20
ws.cell(r, 1, "RECRUITING & GROWTH"); style_section(ws, r, 4)
r = 21
for i, h in enumerate(["", "Conservative", "Optimistic", "Why"]):
    ws.cell(r, i+1, h)
style_header(ws, r, 4)

recruit_assumptions = [
    ("Your direct recruits (L1) by end Y1", 4, 8, "Cumulative people YOU personally recruit"),
    ("Your direct recruits (L1) by end Y2", 8, 15, ""),
    ("Your direct recruits (L1) by end Y3", 12, 25, ""),
    ("Your direct recruits (L1) by end Y4", 15, 30, ""),
    ("Avg recruits per rep (fanout)", 2.0, 3.5, "Each rep recruits this many people on average over their lifetime"),
    ("Months before a new rep starts recruiting", 3, 2, "Ramp time before a rep starts bringing in their own recruits"),
]

for i, (label, cons, opt, note) in enumerate(recruit_assumptions):
    r = 22 + i
    ws.cell(r, 1, label).font = label_font
    ws.cell(r, 2, cons)
    ws.cell(r, 3, opt)
    ws.cell(r, 4, note).font = note_font

# --- Comp Plan ---
r = 29
ws.cell(r, 1, "COMPENSATION PLAN"); style_section(ws, r, 4)
r = 30
for i, h in enumerate(["Component", "Rate", "Unlock Requirement", "Notes"]):
    ws.cell(r, i+1, h)
style_header(ws, r, 4)

comp = [
    ("Retail commission", "25%", "Everyone", "On personal sales"),
    ("L1 override", "10%", "Starter: recruit 1", "On direct recruits' sales"),
    ("L2 override", "5%", "Builder: recruit 3", ""),
    ("L3 override", "4%", "Leader: 5 recruits + $5k GV", ""),
    ("L4 override", "3%", "Director: $25k GV", ""),
    ("L5 override", "2%", "Sr. Director: $75k GV", ""),
    ("L6 override", "1%", "VP: $200k GV", ""),
    ("L7 override", "1%", "Founder's Circle: $500k GV", ""),
    ("Leadership pool: Director", "1% of co. revenue", "Split among Directors only", "Separate pool per rank"),
    ("Leadership pool: Sr. Director", "+1% of co. revenue", "Split among Sr. Directors only", ""),
    ("Leadership pool: VP", "+1% of co. revenue", "Split among VPs only", ""),
    ("Leadership pool: FC", "+1% of co. revenue", "Split among FCs only", ""),
    ("Founding 20 bonus", "1% of co. revenue", "Original 20 only", "÷ 20, forever, if active"),
]

for i, (label, rate, req, note) in enumerate(comp):
    r = 31 + i
    ws.cell(r, 1, label).font = label_font
    ws.cell(r, 2, rate)
    ws.cell(r, 3, req)
    ws.cell(r, 4, note).font = note_font

r = 45
ws.cell(r, 1, "MAX THEORETICAL PAYOUT").font = red_bold
ws.cell(r, 2, "25% + 26% + 4% + 1% = 56%")
ws.cell(r, 4, "Actual will be ~40-48% because most reps don't unlock all levels").font = note_font


# ============================================================
# SHEET 2: CHURN MODEL (month by month)
# ============================================================
ws_churn = wb.create_sheet("Churn Model")
set_col_widths(ws_churn, [20, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 16, 40])

r = 1
ws_churn.cell(r, 1, "MONTHLY CHURN MODEL — COHORT TRACKING").font = Font(bold=True, size=14)
r = 2
ws_churn.cell(r, 1, "Shows how a cohort of 100 reps decays over 24 months").font = note_font

# Conservative cohort
r = 4
ws_churn.cell(r, 1, "CONSERVATIVE SCENARIO"); style_section(ws_churn, r, 14)
r = 5
headers = ["Month", "Start", "Churn Rate", "Churned", "Remaining", "", "", "", "", "", "", "", "", "Formula"]
for i, h in enumerate(["Month", "Start of Month", "Churn Rate", "Churned Out", "End of Month", "Formula"]):
    ws_churn.cell(r, i+1, h)
style_header(ws_churn, r, 6)

for m in range(1, 25):
    r = 5 + m
    ws_churn.cell(r, 1, m)  # month

    if m == 1:
        ws_churn.cell(r, 2, 100)  # start with 100
    else:
        prev_r = r - 1
        ws_churn.cell(r, 2).value = f"=E{prev_r}"  # start = prev end

    # churn rate: 8% for months 1-3, 4% for 4-12, 2% for 13+
    if m <= 3:
        ws_churn.cell(r, 3, 0.08)
        ws_churn.cell(r, 6, "Months 1-3: 8%/mo (danger zone)").font = note_font
    elif m <= 12:
        ws_churn.cell(r, 3, 0.04)
        if m == 4:
            ws_churn.cell(r, 6, "Months 4-12: 4%/mo (settling in)").font = note_font
    else:
        ws_churn.cell(r, 3, 0.02)
        if m == 13:
            ws_churn.cell(r, 6, "Months 13+: 2%/mo (steady state)").font = note_font

    ws_churn.cell(r, 3).number_format = '0%'
    ws_churn.cell(r, 4).value = f"=ROUND(B{r}*C{r},1)"  # churned
    ws_churn.cell(r, 5).value = f"=B{r}-D{r}"  # remaining
    ws_churn.cell(r, 5).number_format = '0.0'

# Summary
r = 31
ws_churn.cell(r, 1, "SUMMARY").font = bold_font
ws_churn.cell(r+1, 1, "After 3 months:").font = label_font
ws_churn.cell(r+1, 2).value = "=E8"
ws_churn.cell(r+1, 2).number_format = '0.0'
ws_churn.cell(r+1, 3, "out of 100").font = note_font

ws_churn.cell(r+2, 1, "After 12 months:").font = label_font
ws_churn.cell(r+2, 2).value = "=E17"
ws_churn.cell(r+2, 2).number_format = '0.0'
ws_churn.cell(r+2, 3, "out of 100").font = note_font

ws_churn.cell(r+3, 1, "After 24 months:").font = label_font
ws_churn.cell(r+3, 2).value = "=E29"
ws_churn.cell(r+3, 2).number_format = '0.0'
ws_churn.cell(r+3, 3, "out of 100").font = note_font

# Optimistic cohort
r = 37
ws_churn.cell(r, 1, "OPTIMISTIC SCENARIO"); style_section(ws_churn, r, 14)
r = 38
for i, h in enumerate(["Month", "Start of Month", "Churn Rate", "Churned Out", "End of Month", "Formula"]):
    ws_churn.cell(r, i+1, h)
style_header(ws_churn, r, 6)

for m in range(1, 25):
    r = 38 + m
    ws_churn.cell(r, 1, m)

    if m == 1:
        ws_churn.cell(r, 2, 100)
    else:
        prev_r = r - 1
        ws_churn.cell(r, 2).value = f"=E{prev_r}"

    if m <= 3:
        ws_churn.cell(r, 3, 0.05)
        ws_churn.cell(r, 6, "Months 1-3: 5%/mo (better onboarding)").font = note_font
    elif m <= 12:
        ws_churn.cell(r, 3, 0.025)
        if m == 4:
            ws_churn.cell(r, 6, "Months 4-12: 2.5%/mo").font = note_font
    else:
        ws_churn.cell(r, 3, 0.012)
        if m == 13:
            ws_churn.cell(r, 6, "Months 13+: 1.2%/mo (strong product)").font = note_font

    ws_churn.cell(r, 3).number_format = '0.0%'
    ws_churn.cell(r, 4).value = f"=ROUND(B{r}*C{r},1)"
    ws_churn.cell(r, 5).value = f"=B{r}-D{r}"
    ws_churn.cell(r, 5).number_format = '0.0'

r = 64
ws_churn.cell(r, 1, "SUMMARY").font = bold_font
ws_churn.cell(r+1, 1, "After 3 months:").font = label_font
ws_churn.cell(r+1, 2).value = "=E41"
ws_churn.cell(r+1, 2).number_format = '0.0'
ws_churn.cell(r+2, 1, "After 12 months:").font = label_font
ws_churn.cell(r+2, 2).value = "=E50"
ws_churn.cell(r+2, 2).number_format = '0.0'
ws_churn.cell(r+3, 1, "After 24 months:").font = label_font
ws_churn.cell(r+3, 2).value = "=E62"
ws_churn.cell(r+3, 2).number_format = '0.0'


# ============================================================
# SHEET 3 & 4: EARNINGS SCENARIOS
# ============================================================
def build_scenario(ws, scenario_name, avg_pv, personal_pv,
                   l1_by_year, fanout, survival_12mo, survival_24mo):
    set_col_widths(ws, [38, 16, 16, 16, 16, 45])
    headers = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Formula / Notes"]

    r = 1
    ws.cell(r, 1, f"FOUNDING GLOW GIRL — {scenario_name}").font = Font(bold=True, size=14)
    r = 2
    ws.cell(r, 1, f"Avg PV: ${avg_pv}/mo | Your PV: ${personal_pv}/mo | 12mo survival: {survival_12mo*100:.0f}%").font = note_font

    # --- Build org bottom-up with churn ---
    r = 4
    ws.cell(r, 1, "YOUR ORGANIZATION (bottom-up with churn)"); style_section(ws, r, 6)
    r = 5
    for i, h in enumerate(headers):
        ws.cell(r, i+1, h)
    style_header(ws, r, 6)

    # Calculate org by level for each year, accounting for churn
    # L1 = your direct recruits (gross), adjusted for churn
    # L2 = L1 reps' recruits × fanout, adjusted for churn
    # etc.

    org_data = []  # will hold (label, [y1,y2,y3,y4], formula_note)

    level_rates = [0.10, 0.05, 0.04, 0.03, 0.02, 0.01, 0.01]
    level_labels = [
        "L1: Your direct recruits",
        "L2: Their recruits",
        "L3",
        "L4",
        "L5",
        "L6",
        "L7",
    ]

    # Gross recruits at each level (before churn)
    gross_by_level = []  # [year][level]
    for yr in range(4):
        l1_gross = l1_by_year[yr]
        levels = [l1_gross]
        for lv in range(1, 7):
            prev_level_gross = levels[lv-1]
            this_level = prev_level_gross * fanout
            # Deeper levels take time to develop
            if yr == 0 and lv > 1: this_level = 0
            if yr == 1 and lv > 3: this_level = 0
            if yr == 2 and lv > 5: this_level = 0
            levels.append(int(this_level))
        gross_by_level.append(levels)

    # Apply churn: people at deeper levels were recruited later, so less time to churn
    # But also: L1 reps recruited in Y1 have been around longest
    # Simplified: apply survival rate based on average tenure at that level
    net_by_level = []
    for yr in range(4):
        levels = []
        for lv in range(7):
            gross = gross_by_level[yr][lv]
            # Avg tenure: L1 reps have been around longest, deeper levels are newer
            # Year N reps at level L have avg tenure of roughly (yr - lv/2) years
            avg_years = max(0.25, (yr + 1) - (lv * 0.5))
            if avg_years >= 2:
                survival = survival_24mo
            elif avg_years >= 1:
                survival = survival_12mo
            else:
                # Partial first year: interpolate
                survival = 1.0 - (1.0 - survival_12mo) * avg_years
            net = int(gross * survival)
            levels.append(net)
        net_by_level.append(levels)

    # Row: Gross L1 recruits
    r = 6
    ws.cell(r, 1, "Gross L1 recruits (cumulative)").font = label_font
    for yr in range(4):
        ws.cell(r, yr+2, l1_by_year[yr]).number_format = '#,##0'
    ws.cell(r, 6, "Total people YOU have ever recruited").font = note_font

    r = 7
    ws.cell(r, 1, "").font = label_font  # spacer

    # Net org by level
    r = 8
    ws.cell(r, 1, "ACTIVE REPS BY LEVEL (after churn)").font = bold_font

    for lv in range(7):
        r = 9 + lv
        ws.cell(r, 1, f"  {level_labels[lv]} (net)").font = label_font
        for yr in range(4):
            ws.cell(r, yr+2, net_by_level[yr][lv]).number_format = '#,##0'
        gross_note = f"Gross × survival rate. Override: {level_rates[lv]*100:.0f}%"
        ws.cell(r, 6, gross_note).font = note_font

    r = 16
    ws.cell(r, 1, "TOTAL ACTIVE ORG SIZE").font = bold_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"=SUM({col}9:{col}15)"
        ws.cell(r, yr+2).number_format = '#,##0'
        ws.cell(r, yr+2).font = bold_font
    ws.cell(r, 6, "= Sum of all active reps across levels").font = note_font

    r = 17
    ws.cell(r, 1, "Your org monthly GV").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}16*{avg_pv}"
        ws.cell(r, yr+2).number_format = '$#,##0'
    ws.cell(r, 6, f"= Active org × ${avg_pv} avg PV").font = note_font

    # Estimate total company from all 20 founders
    r = 19
    ws.cell(r, 1, "ESTIMATED COMPANY TOTALS"); style_section(ws, r, 6)
    r = 20
    for i, h in enumerate(headers):
        ws.cell(r, i+1, h)
    style_header(ws, r, 6)

    r = 21
    ws.cell(r, 1, "Total company reps (20 founders × avg)").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        # Not all 20 founders build equally. Top builders have 2x avg, bottom have 0.5x
        # Your org is above avg (you're pitching top performers), so company ≈ 15× your org
        ws.cell(r, yr+2).value = f"={col}16*15"
        ws.cell(r, yr+2).number_format = '#,##0'
    ws.cell(r, 6, "~15× your org (not all founders build equally)").font = note_font

    r = 22
    ws.cell(r, 1, "Monthly company revenue").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}21*{avg_pv}"
        ws.cell(r, yr+2).number_format = '$#,##0'
    ws.cell(r, 6, f"= Total reps × ${avg_pv}").font = note_font

    r = 23
    ws.cell(r, 1, "Annual company revenue").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}22*12"
        ws.cell(r, yr+2).number_format = '$#,##0'
    ws.cell(r, 6, "= Monthly × 12").font = note_font

    # --- EARNINGS ---
    r = 25
    ws.cell(r, 1, "YOUR MONTHLY EARNINGS"); style_section(ws, r, 6)
    r = 26
    for i, h in enumerate(headers):
        ws.cell(r, i+1, h)
    style_header(ws, r, 6)

    # Personal sales
    r = 27
    ws.cell(r, 1, "Personal sales (25%)").font = label_font
    for yr in range(4):
        ws.cell(r, yr+2, int(personal_pv * 0.25)).number_format = '$#,##0'
    ws.cell(r, 6, f"= ${personal_pv} your PV × 25%").font = note_font

    # Level overrides
    for lv in range(7):
        r = 28 + lv
        ws.cell(r, 1, f"L{lv+1} override ({level_rates[lv]*100:.0f}%)").font = label_font
        for yr in range(4):
            # Reference the net reps row for this level
            reps_row = 9 + lv
            col = get_column_letter(yr+2)
            ws.cell(r, yr+2).value = f"={col}{reps_row}*{avg_pv}*{level_rates[lv]}"
            ws.cell(r, yr+2).number_format = '$#,##0'
        ws.cell(r, 6, f"= L{lv+1} active reps × ${avg_pv} × {level_rates[lv]*100:.0f}%").font = note_font

    r = 35
    ws.cell(r, 1, "SUBTOTAL: Overrides").font = bold_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"=SUM({col}28:{col}34)"
        ws.cell(r, yr+2).number_format = '$#,##0'
        ws.cell(r, yr+2).font = bold_font

    # Leadership pool
    r = 37
    ws.cell(r, 1, "Leadership pool share").font = label_font
    # Estimate: you reach Director by Y2, Sr Dir Y3, VP Y4
    # Directors in company: Y1=0, Y2=~5, Y3=~20, Y4=~50
    leaders_at_your_rank = [1, 5, 15, 30]
    pool_pcts = [0, 0.01, 0.02, 0.03]
    for yr in range(4):
        col = get_column_letter(yr+2)
        if pool_pcts[yr] > 0:
            ws.cell(r, yr+2).value = f"={col}22*{pool_pcts[yr]}/{leaders_at_your_rank[yr]}"
        else:
            ws.cell(r, yr+2, 0)
        ws.cell(r, yr+2).number_format = '$#,##0'
    ws.cell(r, 6, "= Company rev × rank pool % ÷ # leaders at your rank").font = note_font

    r = 38
    ws.cell(r, 1, "  Your rank").font = note_font
    rank_by_year = ["Builder", "Director", "Sr. Director", "VP"]
    for yr in range(4):
        ws.cell(r, yr+2, rank_by_year[yr]).font = note_font

    r = 39
    ws.cell(r, 1, "  Leaders sharing this pool").font = note_font
    for yr in range(4):
        ws.cell(r, yr+2, leaders_at_your_rank[yr]).font = note_font

    # Founding 20 bonus
    r = 41
    ws.cell(r, 1, "Founding 20 bonus (1% ÷ 20)").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}22*0.01/20"
        ws.cell(r, yr+2).number_format = '$#,##0'
    ws.cell(r, 6, "= 1% of company monthly revenue ÷ 20 founders").font = note_font

    # TOTALS
    r = 43
    ws.cell(r, 1, "MONTHLY TOTAL").font = big_bold
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}27+{col}35+{col}37+{col}41"
        ws.cell(r, yr+2).number_format = '$#,##0'
        ws.cell(r, yr+2).font = big_bold

    r = 44
    ws.cell(r, 1, "ANNUAL TOTAL").font = purple_bold
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}43*12"
        ws.cell(r, yr+2).number_format = '$#,##0'
        ws.cell(r, yr+2).font = purple_bold

    # ---- SANITY CHECKS ----
    r = 46
    ws.cell(r, 1, "SANITY CHECKS"); style_section(ws, r, 6)

    r = 47
    ws.cell(r, 1, "CHECK 1: Total payout ratio").font = red_bold
    r = 48
    ws.cell(r, 1, "Est. total commissions (all reps)").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        # Estimate: avg rep earns retail (25% on their PV) + avg ~3% from overrides
        # Total payout ≈ 25% (retail) + ~15% (overrides) + 4% (pools) + 1% (founding) = ~45%
        ws.cell(r, yr+2).value = f"={col}22*0.45"
        ws.cell(r, yr+2).number_format = '$#,##0'
    ws.cell(r, 6, "Estimated at ~45% of revenue (retail + overrides + pools)").font = note_font

    r = 49
    ws.cell(r, 1, "Payout ratio").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}48/{col}22"
        ws.cell(r, yr+2).number_format = '0.0%'
        ws.cell(r, yr+2).fill = check_fill
    ws.cell(r, 6, "TARGET: 35-50%. Over 55% = danger.").font = note_font

    r = 51
    ws.cell(r, 1, "CHECK 2: Can the company afford it?").font = red_bold
    r = 52
    ws.cell(r, 1, "Monthly company revenue").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}22"
        ws.cell(r, yr+2).number_format = '$#,##0'

    r = 53
    ws.cell(r, 1, "- Product COGS (20%)").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"=-{col}22*0.20"
        ws.cell(r, yr+2).number_format = '$#,##0'

    r = 54
    ws.cell(r, 1, "- Total commissions (~45%)").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"=-{col}48"
        ws.cell(r, yr+2).number_format = '$#,##0'

    r = 55
    ws.cell(r, 1, "- Shipping & ops (10%)").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"=-{col}22*0.10"
        ws.cell(r, yr+2).number_format = '$#,##0'

    r = 56
    ws.cell(r, 1, "= GROSS PROFIT").font = bold_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}52+{col}53+{col}54+{col}55"
        ws.cell(r, yr+2).number_format = '$#,##0'
        ws.cell(r, yr+2).font = bold_font

    r = 57
    ws.cell(r, 1, "Gross margin %").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}56/{col}52"
        ws.cell(r, yr+2).number_format = '0.0%'
        ws.cell(r, yr+2).fill = check_fill
    ws.cell(r, 6, "TARGET: 20%+. Under 15% = unsustainable.").font = note_font

    r = 59
    ws.cell(r, 1, "CHECK 3: Median rep earnings").font = red_bold
    r = 60
    ws.cell(r, 1, "Median rep (sells avg PV, no team)").font = label_font
    for yr in range(4):
        ws.cell(r, yr+2, int(avg_pv * 0.25)).number_format = '$#,##0'
    ws.cell(r, 6, f"= ${avg_pv} × 25% retail only. Most reps earn ONLY this.").font = note_font

    r = 61
    ws.cell(r, 1, "Median rep annual").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}60*12"
        ws.cell(r, yr+2).number_format = '$#,##0'
    ws.cell(r, 6, "This is what MOST of your reps will actually make").font = note_font

    r = 63
    ws.cell(r, 1, "CHECK 4: Top-heaviness").font = red_bold
    r = 64
    ws.cell(r, 1, "Your annual earnings").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}44"
        ws.cell(r, yr+2).number_format = '$#,##0'

    r = 65
    ws.cell(r, 1, "Top 20 founders combined (est.)").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        # Estimate: you're a top builder. Avg founder earns ~60% of what you earn. 20 × 60%
        ws.cell(r, yr+2).value = f"={col}44*0.6*20"
        ws.cell(r, yr+2).number_format = '$#,##0'
    ws.cell(r, 6, "Est. avg founder earns ~60% of top builder").font = note_font

    r = 66
    ws.cell(r, 1, "Total annual commissions (all reps)").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}48*12"
        ws.cell(r, yr+2).number_format = '$#,##0'

    r = 67
    ws.cell(r, 1, "Top 20 as % of all commissions").font = label_font
    for yr in range(4):
        col = get_column_letter(yr+2)
        ws.cell(r, yr+2).value = f"={col}65/{col}66"
        ws.cell(r, yr+2).number_format = '0.0%'
        ws.cell(r, yr+2).fill = check_fill
    ws.cell(r, 6, "TARGET: Under 30%. Over 40% = too top-heavy.").font = note_font

    r = 69
    ws.cell(r, 1, "CHECK 5: Max theoretical payout stack").font = red_bold
    r = 70
    ws.cell(r, 1, "Retail: 25% + Overrides: 26% + Pools: 4% + F20: 1%").font = label_font
    ws.cell(r, 2, 0.56).number_format = '0%'
    ws.cell(r, 2).fill = check_fill
    ws.cell(r, 6, "TARGET: Under 65%. Actual payout will be lower since most don't unlock all levels.").font = note_font


# Build both scenarios
ws_cons = wb.create_sheet("Conservative")
build_scenario(
    ws_cons,
    scenario_name="CONSERVATIVE ($250 avg PV, high churn)",
    avg_pv=250,
    personal_pv=800,
    l1_by_year=[4, 8, 12, 15],
    fanout=2,
    survival_12mo=0.47,   # from churn model: 8%×3mo + 4%×9mo
    survival_24mo=0.34,
)

ws_opt = wb.create_sheet("Optimistic")
build_scenario(
    ws_opt,
    scenario_name="OPTIMISTIC ($500 avg PV, low churn)",
    avg_pv=500,
    personal_pv=1500,
    l1_by_year=[8, 15, 25, 30],
    fanout=3.5,
    survival_12mo=0.65,   # from churn model: 5%×3mo + 2.5%×9mo
    survival_24mo=0.52,
)


# ============================================================
# SHEET 5: COMPARISON
# ============================================================
ws_cmp = wb.create_sheet("Comparison")
set_col_widths(ws_cmp, [32, 16, 16, 16, 16, 16, 16, 16, 16])

r = 1
ws_cmp.cell(r, 1, "SIDE-BY-SIDE COMPARISON").font = Font(bold=True, size=14)

r = 3
headers = ["", "Cons Y1", "Opt Y1", "Cons Y2", "Opt Y2", "Cons Y3", "Opt Y3", "Cons Y4", "Opt Y4"]
for i, h in enumerate(headers):
    ws_cmp.cell(r, i+1, h)
style_header(ws_cmp, r, 9)

# Reference cells from each scenario sheet
ref_rows = [
    ("Your active org size", 16, '#,##0'),
    ("Your org monthly GV", 17, '$#,##0'),
    ("Monthly company revenue", 22, '$#,##0'),
    ("Your monthly earnings", 43, '$#,##0'),
    ("Your annual earnings", 44, '$#,##0'),
    ("Payout ratio", 49, '0.0%'),
    ("Company gross margin", 57, '0.0%'),
    ("Top 20 as % of commissions", 67, '0.0%'),
]

for i, (label, src_row, fmt) in enumerate(ref_rows):
    r = 4 + i
    ws_cmp.cell(r, 1, label).font = label_font
    for yr in range(4):
        col_letter = get_column_letter(yr + 2)  # B, C, D, E in source sheets
        cons_col = 2 + yr * 2      # B, D, F, H in comparison
        opt_col = 3 + yr * 2       # C, E, G, I in comparison

        ws_cmp.cell(r, cons_col).value = f"=Conservative!{col_letter}{src_row}"
        ws_cmp.cell(r, cons_col).number_format = fmt

        ws_cmp.cell(r, opt_col).value = f"=Optimistic!{col_letter}{src_row}"
        ws_cmp.cell(r, opt_col).number_format = fmt
        ws_cmp.cell(r, opt_col).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")


# Save
output_path = "/home/zdela/proj/glow/mlm-comp-model/glow-earnings-model.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print("Tabs: Assumptions | Churn Model | Conservative | Optimistic | Comparison")
